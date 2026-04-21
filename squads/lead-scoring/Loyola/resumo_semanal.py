#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Script para gerar resumo semanal de campanhas de marketing
Lê dados de arquivo Excel exportado de Google Sheets e produz texto formatado para ClickUp

Requisitos:
  - openpyxl: pip install openpyxl

Uso:
  python3 resumo_semanal.py arquivo.xlsx 19/02 25/02 [--historico historico.json]
"""

import argparse
import json
import sys
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional, Tuple
import re
from difflib import SequenceMatcher

try:
    from openpyxl import load_workbook
except ImportError:
    print("Erro: openpyxl não está instalado.")
    print("Execute: pip install openpyxl")
    sys.exit(1)


class ResumoSemanal:
    """Processa dados de Excel e gera resumo semanal de campanhas"""

    def __init__(self, arquivo_xlsx: str, data_inicio: str, data_fim: str,
                 arquivo_historico: Optional[str] = None):
        """
        Inicializa o processador

        Args:
            arquivo_xlsx: Caminho do arquivo .xlsx
            data_inicio: Data inicial no formato DD/MM/YYYY
            data_fim: Data final no formato DD/MM/YYYY
            arquivo_historico: Caminho opcional do arquivo JSON com histórico
        """
        self.arquivo_xlsx = arquivo_xlsx
        self.data_inicio = self._parse_data(data_inicio)
        self.data_fim = self._parse_data(data_fim)
        self.historico = self._carregar_historico(arquivo_historico)

        # Carrega workbook com otimizações para arquivo grande
        try:
            self.workbook = load_workbook(arquivo_xlsx, read_only=True, data_only=True)
        except Exception as e:
            print(f"Erro ao abrir arquivo: {e}")
            sys.exit(1)

        # Dados extraídos
        self.dados_macro = {}
        self.dados_lp_test = {}
        self.dados_ads_hot = {}
        self.dados_ads_cold = {}
        self.dados_pesquisa = {}

    @staticmethod
    def _parse_data(data_str: str) -> datetime:
        """Converte string DD/MM/YYYY para datetime"""
        try:
            return datetime.strptime(data_str, "%d/%m/%Y")
        except ValueError:
            print(f"Erro: Formato de data inválido: {data_str}. Use DD/MM/YYYY")
            sys.exit(1)

    def _carregar_historico(self, arquivo_historico: Optional[str]) -> Dict:
        """Carrega arquivo JSON com dados históricos"""
        if not arquivo_historico:
            return {}
        try:
            with open(arquivo_historico, 'r', encoding='utf-8') as f:
                return json.load(f)
        except Exception as e:
            print(f"Aviso: Erro ao carregar histórico: {e}")
            return {}

    def _encontrar_sheet(self, nome_esperado: str) -> Optional:
        """
        Encontra sheet com fuzzy matching (ignora espaços e emojis)

        Args:
            nome_esperado: Nome da sheet a encontrar

        Returns:
            Sheet encontrada ou None
        """
        # Remove emojis e espaços extras
        nome_limpo = self._limpar_nome(nome_esperado)

        for sheet_name in self.workbook.sheetnames:
            sheet_limpa = self._limpar_nome(sheet_name)
            if sheet_limpa == nome_limpo:
                return self.workbook[sheet_name]

        # Se não encontrar exatamente, tenta fuzzy matching
        for sheet_name in self.workbook.sheetnames:
            similaridade = SequenceMatcher(None,
                                          self._limpar_nome(sheet_name),
                                          nome_limpo).ratio()
            if similaridade > 0.8:
                return self.workbook[sheet_name]

        return None

    @staticmethod
    def _limpar_nome(nome: str) -> str:
        """Remove emojis, espaços extras e caracteres especiais"""
        # Remove emojis (bloco Unicode)
        nome = re.sub(r'[\U0001F300-\U0001F9FF]', '', nome)
        # Remove espaços extras
        nome = ' '.join(nome.split())
        return nome.strip()

    @staticmethod
    def _safe_float(valor) -> float:
        """Converte valor para float com segurança"""
        if valor is None:
            return 0.0
        try:
            return float(valor)
        except (ValueError, TypeError):
            return 0.0

    @staticmethod
    def _formatar_moeda(valor: float, casas_decimais: int = 1) -> str:
        """Formata valor em reais brasileiros"""
        # Arredonda
        valor = round(valor, casas_decimais)

        # Formata com separador de milhares
        if valor >= 100:
            formatado = f"{valor:,.0f}".replace(",", ".")
        else:
            formatado = f"{valor:.{casas_decimais}f}".replace(".", ",")

        return f"R$ {formatado}"

    @staticmethod
    def _formatar_percentual(valor: float, casas_decimais: int = 1) -> str:
        """Formata percentual em formato brasileiro"""
        valor_pct = round(valor * 100, casas_decimais)
        formatado = f"{valor_pct:.{casas_decimais}f}".replace(".", ",")
        return f"{formatado}%"

    @staticmethod
    def _formatar_pct_simples(valor: float) -> float:
        """Retorna percentual como float (multiplicado por 100)"""
        return round(valor * 100, 2)

    def extrair_dados_macro(self):
        """Extrai dados da aba 'Captação - Macro 2026'"""
        ws = self._encontrar_sheet("Captação - Macro 2026")
        if not ws:
            print("Aviso: Aba 'Captação - Macro 2026' não encontrada")
            return

        self.dados_macro = {
            "investimento": 0,
            "cliques": 0,
            "impressoes": 0,
            "cpm": [],
            "cpc": [],
            "ctr": [],
            "lp_view": 0,
            "connect_rate": [],
            "tx_conv": [],
            "leads_pagos": 0,
            "leads_org": 0,
            "leads_sem_track": 0,
            "cpl_pago": 0,
            "pesquisa": 0,
            "pesquisa_sem_email": 0,
            "tx_resposta": 0,
        }

        # Itera pelas linhas
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row[0]:  # Se não houver data, pula
                continue

            try:
                # Coluna A: Data
                if isinstance(row[0], datetime):
                    data = row[0]
                else:
                    continue

                # Verifica se está no intervalo
                if not (self.data_inicio <= data <= self.data_fim):
                    continue

                # Soma investimento (coluna B)
                self.dados_macro["investimento"] += self._safe_float(row[1])

                # Soma cliques (coluna C)
                self.dados_macro["cliques"] += self._safe_float(row[2])

                # Soma impressões (coluna D)
                self.dados_macro["impressoes"] += self._safe_float(row[3])

                # Armazena CPM, CPC, CTR, Connect Rate (para depois calcular médias)
                if row[4]:  # CPM
                    self.dados_macro["cpm"].append(self._safe_float(row[4]))
                if row[5]:  # CPC
                    self.dados_macro["cpc"].append(self._safe_float(row[5]))
                if row[6]:  # CTR (já em decimal)
                    self.dados_macro["ctr"].append(self._safe_float(row[6]))

                # Soma LP View (coluna H)
                self.dados_macro["lp_view"] += self._safe_float(row[7])

                # Connect Rate (coluna I) - pode ser None, será calculado
                if row[8]:
                    self.dados_macro["connect_rate"].append(self._safe_float(row[8]))
                else:
                    # Calcula: LP View / Cliques
                    cliques = self._safe_float(row[2])
                    lp_view = self._safe_float(row[7])
                    if cliques > 0:
                        self.dados_macro["connect_rate"].append(lp_view / cliques)

                # Tx Conv (coluna J)
                if row[9]:
                    self.dados_macro["tx_conv"].append(self._safe_float(row[9]))

                # Leads pagos (coluna K)
                self.dados_macro["leads_pagos"] += self._safe_float(row[10])

                # Leads org (coluna L)
                self.dados_macro["leads_org"] += self._safe_float(row[11])

                # Leads s/track (coluna M)
                self.dados_macro["leads_sem_track"] += self._safe_float(row[12])

                # Pesquisa (coluna P)
                self.dados_macro["pesquisa"] += self._safe_float(row[15])

                # Pesquisa sem e-mail (coluna Q)
                self.dados_macro["pesquisa_sem_email"] += self._safe_float(row[16])

                # Tx Resposta (coluna R)
                if row[17]:
                    self.dados_macro["tx_resposta"] = max(self._safe_float(row[17]),
                                                         self.dados_macro["tx_resposta"])

            except (IndexError, TypeError):
                continue

        # Calcula CPL pago
        leads_pagos = self.dados_macro["leads_pagos"]
        if leads_pagos > 0:
            self.dados_macro["cpl_pago"] = self.dados_macro["investimento"] / leads_pagos

        # Calcula médias
        if self.dados_macro["cpm"]:
            self.dados_macro["cpm_media"] = sum(self.dados_macro["cpm"]) / len(self.dados_macro["cpm"])
        if self.dados_macro["cpc"]:
            self.dados_macro["cpc_media"] = sum(self.dados_macro["cpc"]) / len(self.dados_macro["cpc"])
        if self.dados_macro["ctr"]:
            self.dados_macro["ctr_media"] = sum(self.dados_macro["ctr"]) / len(self.dados_macro["ctr"])
        if self.dados_macro["connect_rate"]:
            self.dados_macro["connect_rate_media"] = (
                sum(self.dados_macro["connect_rate"]) / len(self.dados_macro["connect_rate"])
            )

    def extrair_dados_lp_test(self):
        """Extrai dados da aba 'Captação - Teste LP'"""
        ws = self._encontrar_sheet("Captação - Teste LP")
        if not ws:
            print("Aviso: Aba 'Captação - Teste LP' não encontrada")
            return

        self.dados_lp_test = {}
        lp_atual = None

        # Itera por todas as linhas detectando seções LPA/LPB
        for row in ws.iter_rows(values_only=True):
            if not row or not row[0]:
                continue

            cell_val = str(row[0]).strip().upper()

            # Detecta marcador de nova LP (LPA, LPB, LP01, LP02, etc.)
            if cell_val in ("LPA", "LPB", "LPC", "LPD") or \
               (cell_val.startswith("LP") and len(cell_val) <= 5 and cell_val != "LP VIEW"):
                lp_atual = cell_val
                if lp_atual not in self.dados_lp_test:
                    self.dados_lp_test[lp_atual] = {
                        "investimento": 0, "cliques": 0, "impressoes": 0,
                        "lp_view": 0, "leads": 0, "connect_rate": [],
                        "tx_conv": [], "cpl": 0, "data_inicio": None,
                    }
                continue

            # Pula linhas de cabeçalho
            if cell_val == "DIA" or not isinstance(row[0], datetime):
                continue

            # Se não temos LP atual definida, pula
            if not lp_atual:
                continue

            try:
                data = row[0]
                if not (self.data_inicio <= data <= self.data_fim):
                    continue

                # Registra data de início
                if not self.dados_lp_test[lp_atual]["data_inicio"]:
                    self.dados_lp_test[lp_atual]["data_inicio"] = data

                self.dados_lp_test[lp_atual]["investimento"] += self._safe_float(row[1])
                self.dados_lp_test[lp_atual]["cliques"] += self._safe_float(row[2])
                self.dados_lp_test[lp_atual]["impressoes"] += self._safe_float(row[3])
                self.dados_lp_test[lp_atual]["lp_view"] += self._safe_float(row[7])
                self.dados_lp_test[lp_atual]["leads"] += self._safe_float(row[10])

                # Connect Rate
                if len(row) > 8 and row[8]:
                    self.dados_lp_test[lp_atual]["connect_rate"].append(self._safe_float(row[8]))
                else:
                    cliques = self._safe_float(row[2])
                    lp_view = self._safe_float(row[7])
                    if cliques > 0:
                        self.dados_lp_test[lp_atual]["connect_rate"].append(lp_view / cliques)

                # Tx Conv
                if len(row) > 9 and row[9]:
                    self.dados_lp_test[lp_atual]["tx_conv"].append(self._safe_float(row[9]))

            except (IndexError, TypeError):
                continue

        # Calcula médias para cada LP
        for lp_nome in self.dados_lp_test:
            lp = self.dados_lp_test[lp_nome]
            if lp["connect_rate"]:
                lp["connect_rate_media"] = sum(lp["connect_rate"]) / len(lp["connect_rate"])
            if lp["tx_conv"]:
                lp["tx_conv_media"] = sum(lp["tx_conv"]) / len(lp["tx_conv"])

    def extrair_dados_ads(self):
        """Extrai dados das abas 'Captação - Ads' (hot/cold)"""
        self.dados_ads_hot = {}
        self.dados_ads_cold = {}

        # Lê Captação - Ads (hot por padrão, pode estar em vendas ou leads)
        ws = self._encontrar_sheet("Captação - Ads")
        if ws:
            self._processar_ads_sheet(ws, self.dados_ads_hot)

        # Lê Ads-Cold se existir
        ws = self._encontrar_sheet("Ads-Cold")
        if ws:
            self._processar_ads_sheet(ws, self.dados_ads_cold)

    def _processar_ads_sheet(self, ws, dados_alvo: Dict):
        """
        Processa sheet de ads e armazena dados

        Args:
            ws: Worksheet a processar
            dados_alvo: Dicionário onde armazenar os dados
        """
        # Primeira linha pode ter info sobre filtros
        # Linha 2 tem headers: Ad Name, Invest, Impressões, Cliques, CTR, CPC, CPM, Leads, CPL, ...

        for row in ws.iter_rows(min_row=3, values_only=True):
            if not row[0]:  # Se não houver nome do ad, pula
                continue

            try:
                ad_name = str(row[0]).strip()
                leads = self._safe_float(row[7]) if len(row) > 7 else 0

                # Ignora linhas de total ou cabeçalho
                if ad_name.lower() in ("total", "ad name", ""):
                    continue

                # Só armazena se tiver leads > 0
                if leads > 0:
                    dados_alvo[ad_name] = {
                        "investimento": self._safe_float(row[1]),
                        "impressoes": self._safe_float(row[2]),
                        "cliques": self._safe_float(row[3]),
                        "ctr": self._safe_float(row[4]),
                        "cpc": self._safe_float(row[5]),
                        "cpm": self._safe_float(row[6]),
                        "leads": leads,
                        "cpl": self._safe_float(row[8]) if len(row) > 8 else 0,
                    }

            except (IndexError, TypeError, ValueError):
                continue

    def extrair_dados_pesquisa(self):
        """Extrai dados da aba 'Resultado-Pesquisa'"""
        ws = self._encontrar_sheet("Resultado-Pesquisa")
        if not ws:
            print("Aviso: Aba 'Resultado-Pesquisa' não encontrada")
            return

        self.dados_pesquisa = {}

        # Lê todas as linhas
        linhas = []
        for row in ws.iter_rows(values_only=True):
            linhas.append(row)

        # Procura pelas seções por cabeçalho (ordem importa!)
        for i, row in enumerate(linhas):
            if not row:
                continue

            # Estado civil (coluna 0) - verificar antes de "civil"
            if row[0] and "estado civil" in str(row[0]).strip().lower():
                if "estado_civil" not in self.dados_pesquisa:
                    self._extrair_secao_pesquisa(linhas, i, "estado_civil", 0, 1, 2)

            # Sexo (coluna 0)
            elif row[0] and "sexo" in str(row[0]).strip().lower():
                if "sexo" not in self.dados_pesquisa:
                    self._extrair_secao_pesquisa(linhas, i, "sexo", 0, 1, 2)

            # Escolaridade (coluna 0) - verificar antes de "idade"
            elif row[0] and "escolaridade" in str(row[0]).strip().lower():
                if "escolaridade" not in self.dados_pesquisa:
                    self._extrair_secao_pesquisa(linhas, i, "escolaridade", 0, 1, 2)

            # Idade (coluna 0)
            elif row[0] and str(row[0]).strip().lower().startswith("e a sua idade"):
                if "idade" not in self.dados_pesquisa:
                    self._extrair_secao_pesquisa(linhas, i, "idade", 0, 1, 2)

            # Filhos (coluna 4)
            if len(row) > 4 and row[4] and "filhos" in str(row[4]).strip().lower():
                if "filhos" not in self.dados_pesquisa:
                    self._extrair_secao_pesquisa(linhas, i, "filhos", 4, 5, None)

            # Religião (coluna 4) - procura por "qual é a sua religião"
            if (len(row) > 4 and row[4] and
                "qual é a sua religião" in str(row[4]).strip().lower()):
                if "religiao" not in self.dados_pesquisa:
                    self._extrair_secao_pesquisa(linhas, i, "religiao", 4, 5, 6)

            # Renda (coluna 4) - procura por "qual é a sua renda"
            if (len(row) > 4 and row[4] and
                "qual é a sua renda" in str(row[4]).strip().lower()):
                if "renda" not in self.dados_pesquisa:
                    self._extrair_secao_pesquisa(linhas, i, "renda", 4, 5, 6)

    def _extrair_secao_pesquisa(self, linhas: List, idx_header: int, nome_secao: str,
                               col_nome: int, col_count: int, col_pct: Optional[int]):
        """
        Extrai dados de uma seção da pesquisa

        Args:
            linhas: Lista de todas as linhas
            idx_header: Índice da linha com o cabeçalho
            nome_secao: Nome da seção (chave do dicionário)
            col_nome: Coluna com nome da opção
            col_count: Coluna com contagem
            col_pct: Coluna com percentual (pode ser None)
        """
        self.dados_pesquisa[nome_secao] = {}

        # Processa próximas 15 linhas (ajustável)
        for i in range(idx_header + 1, min(idx_header + 15, len(linhas))):
            row = linhas[i]
            if not row or len(row) <= col_nome or not row[col_nome]:
                continue

            nome_opcao = str(row[col_nome]).strip()
            if "total geral" in nome_opcao.lower():
                break

            # Pula linhas vazias
            if not nome_opcao or nome_opcao == "None" or (col_count < len(row) and not row[col_count]):
                continue

            try:
                count = int(self._safe_float(row[col_count])) if col_count < len(row) else 0
                pct = self._safe_float(row[col_pct]) if (col_pct is not None and col_pct < len(row)) else 0

                if count > 0:  # Só adiciona se tiver contagem
                    self.dados_pesquisa[nome_secao][nome_opcao] = {
                        "count": count,
                        "percentual": pct
                    }
            except (ValueError, TypeError):
                continue

    def gerar_resumo(self) -> str:
        """Gera o texto do resumo formatado"""
        # Extrai todos os dados
        self.extrair_dados_macro()
        self.extrair_dados_lp_test()
        self.extrair_dados_ads()
        self.extrair_dados_pesquisa()

        linhas = []

        # Cabeçalho
        data_inicio_fmt = self.data_inicio.strftime("%d/%m")
        data_fim_fmt = self.data_fim.strftime("%d/%m")
        linhas.append(f"Resumão {data_inicio_fmt} a {data_fim_fmt}:\n")

        # Seção Geral
        linhas.append("Geral (aba - Captação - Macro 2026)\n")
        if self.dados_macro:
            linhas.append(self._gerar_secao_macro())

        # Seção LP Test
        if self.dados_lp_test:
            linhas.append("\nTeste de LPs (aba - Captação - Teste LP)\n")
            linhas.append(self._gerar_secao_lp_test())

        # Seção Ads
        if self.dados_ads_hot or self.dados_ads_cold:
            linhas.append("\nDesempenho de Criativos (aba - Captação - Ads)\n")
            if self.dados_ads_hot:
                linhas.append(self._gerar_secao_ads("hot"))
            if self.dados_ads_cold:
                linhas.append(self._gerar_secao_ads("cold"))

        # Seção Pesquisa
        if self.dados_pesquisa:
            linhas.append("\nResultado da Pesquisa (aba Resultado-Pesquisa)\n")
            linhas.append(self._gerar_secao_pesquisa())

        return "".join(linhas)

    def _gerar_secao_macro(self) -> str:
        """Gera seção de dados macro"""
        linhas = []

        # Investimento e leads
        inv = self.dados_macro.get("investimento", 0)
        leads_pag = int(self.dados_macro.get("leads_pagos", 0))
        leads_org = int(self.dados_macro.get("leads_org", 0))
        leads_sem = int(self.dados_macro.get("leads_sem_track", 0))

        linhas.append(f"Investimento: {self._formatar_moeda(inv)}\n")
        linhas.append(f"Leads: {leads_pag + leads_org + leads_sem} "
                      f"({leads_pag} pag / {leads_org} org / {leads_sem} sem track)\n")

        # CPL pago
        if leads_pag > 0:
            cpl = inv / leads_pag
            linhas.append(f"CPL pago {self._formatar_moeda(cpl)}\n")

        # Taxa de resposta na pesquisa
        # Os dados de Pesquisa/Tx Resposta podem ser 0 no export (fórmulas não exportam)
        # Nesse caso, calcula a partir do Resultado-Pesquisa e total de leads
        pesquisa = int(self.dados_macro.get("pesquisa", 0))
        pesquisa_sem_email = int(self.dados_macro.get("pesquisa_sem_email", 0))
        total_respondentes = pesquisa + pesquisa_sem_email
        leads_total = leads_pag + leads_org + leads_sem

        if total_respondentes > 0 and leads_total > 0:
            tx_resp = round(total_respondentes / leads_total * 100, 1)
            tx_resp_fmt = f"{tx_resp:.1f}".replace(".", ",")
            linhas.append(f"\nTaxa de resposta na pesquisa: "
                         f"{tx_resp_fmt}% "
                         f"({total_respondentes} → {pesquisa_sem_email} vieram sem e-mail - "
                         f"origem Whatsapp obrigado)\n")
        else:
            # Quando Pesquisa = 0 no export (fórmulas não exportaram),
            # marca para preenchimento manual
            linhas.append(f"\nTaxa de resposta na pesquisa: [PREENCHER - dados de fórmula não disponíveis no export]\n")

        # Connect Rate
        if "connect_rate_media" in self.dados_macro:
            cr = self.dados_macro["connect_rate_media"]
            linhas.append(f"\nConnect Rate: {self._formatar_percentual(cr, 1)}\n")

        # CPM, CTR, CPC com histórico
        linhas.append("\n")
        if "cpm_media" in self.dados_macro:
            cpm = self.dados_macro["cpm_media"]
            linhas.append(f"CPM {self._formatar_moeda(cpm)}")
            if "campanhas_anteriores" in self.historico:
                historico_str = self._gerar_historico_metricas(
                    "CPM", self.historico["campanhas_anteriores"], "CPM"
                )
                linhas.append(historico_str)
            linhas.append("\n")

        if "ctr_media" in self.dados_macro:
            ctr = self.dados_macro["ctr_media"]
            linhas.append(f"CTR {self._formatar_percentual(ctr, 2)}")
            if "campanhas_anteriores" in self.historico:
                historico_str = self._gerar_historico_metricas(
                    "CTR", self.historico["campanhas_anteriores"], "CTR"
                )
                linhas.append(historico_str)
            linhas.append("\n")

        if "cpc_media" in self.dados_macro:
            cpc = self.dados_macro["cpc_media"]
            linhas.append(f"CPC {self._formatar_moeda(cpc)}")
            if "campanhas_anteriores" in self.historico:
                historico_str = self._gerar_historico_metricas(
                    "CPC", self.historico["campanhas_anteriores"], "CPC"
                )
                linhas.append(historico_str)
            linhas.append("\n")

        return "".join(linhas)

    def _gerar_historico_metricas(self, metrica_nome: str, campanhas: Dict,
                                 chave: str) -> str:
        """Gera string com histórico de métricas"""
        if not campanhas:
            return ""
        partes = []
        for campanha_nome, dados in campanhas.items():
            if chave not in dados:
                continue
            valor = dados[chave]
            if isinstance(valor, str):
                partes.append(f"{campanha_nome} foi {valor}")
            else:
                if metrica_nome == "CTR":
                    valor_fmt = f"{valor:.2f}%".replace(".", ",")
                else:
                    valor_fmt = self._formatar_moeda(valor)
                partes.append(f"{campanha_nome} foi {valor_fmt}")
        return " (" + " e ".join(partes) + ")*" if partes else ""

    def _gerar_secao_lp_test(self) -> str:
        """Gera seção de teste de LPs"""
        linhas = []

        # Encontra a primeira data do teste (a data real de início)
        datas_inicio = [d["data_inicio"] for d in self.dados_lp_test.values() if d.get("data_inicio")]
        if datas_inicio:
            primeira_data = min(datas_inicio)
            data_fmt = primeira_data.strftime("%d/%m")
        else:
            data_fmt = self.data_inicio.strftime("%d/%m")

        # Só mostra seção se tiver mais de uma LP (teste A/B)
        if len(self.dados_lp_test) >= 2:
            linhas.append(f"Iniciamos no dia {data_fmt} um teste de LP "
                         f"para o público frio:\n")
        else:
            # Se só tem uma LP, não é teste A/B
            return ""

        # Lista LPs
        for lp_nome, dados in sorted(self.dados_lp_test.items()):
            linhas.append(f"{lp_nome} → (URL a preencher)\n")

        linhas.append("\n")

        # Compara LPs
        lps = list(self.dados_lp_test.items())
        if len(lps) >= 2:
            lpa_nome, lpa_dados = lps[0]
            lpb_nome, lpb_dados = lps[1]
            tx_a = lpa_dados.get("tx_conv_media", 0)
            tx_b = lpb_dados.get("tx_conv_media", 0)
            if tx_a > 0 and tx_b > 0:
                melhor = lpb_nome if tx_b > tx_a else lpa_nome
                linhas.append(f"{melhor} iniciou com uma taxa de conversão maior.\n\n")

        # Connect Rate por LP
        for lp_nome, dados in sorted(self.dados_lp_test.items()):
            if "connect_rate_media" in dados:
                cr = dados["connect_rate_media"]
                linhas.append(f"Connect Rate da {lp_nome} está em "
                             f"{self._formatar_percentual(cr, 1)}\n")

        return "".join(linhas)

    def _gerar_secao_ads(self, tipo: str) -> str:
        """Gera seção de criativos (hot ou cold)"""
        linhas = []
        dados = self.dados_ads_hot if tipo == "hot" else self.dados_ads_cold

        if not dados:
            return ""

        tipo_nome = "quente (hot)" if tipo == "hot" else "frio (cold)"
        linhas.append(f"Público {tipo_nome}\n\n")

        # Separa por tipo de criativo
        estaticos = {}
        videos = {}
        carousels = {}

        for ad_name, dados_ad in dados.items():
            ad_lower = ad_name.lower()
            leads = dados_ad.get("leads", 0)

            if ad_lower.startswith("v"):
                videos[ad_name] = leads
            elif ad_lower.startswith("c"):
                carousels[ad_name] = leads
            else:
                estaticos[ad_name] = leads

        # Melhor estático
        if estaticos:
            melhor_estatico = max(estaticos, key=estaticos.get)
            linhas.append(f"Melhor estático: __{melhor_estatico}__\n")

        # Melhores vídeos (top 2)
        if videos:
            videos_sorted = sorted(videos.items(), key=lambda x: x[1], reverse=True)
            melhor_video = videos_sorted[0][0] if videos_sorted else None
            if len(videos_sorted) > 1:
                segundo_video = videos_sorted[1][0]
                linhas.append(f"Melhores vídeos: __{melhor_video}__ e __{segundo_video}__\n")
            elif melhor_video:
                linhas.append(f"Melhor vídeo: __{melhor_video}__\n")

        # Melhores carousels (se houver)
        if carousels:
            melhor_carousel = max(carousels, key=carousels.get)
            linhas.append(f"Melhor carousel: __{melhor_carousel}__\n")

        linhas.append("\n")
        return "".join(linhas)

    def _gerar_secao_pesquisa(self) -> str:
        """Gera seção de resultado da pesquisa"""
        linhas = []
        medias = self.historico.get("medias_pesquisa", {})

        # Sexo (93% mulheres)
        if "sexo" in self.dados_pesquisa:
            for opcao, dados in self.dados_pesquisa["sexo"].items():
                if "mulher" in opcao.lower():
                    pct = int(round(dados['percentual'] * 100, 0))
                    linhas.append(f"{pct}% mulheres\n")

        # Idade (74% acima de 36 anos)
        if "idade" in self.dados_pesquisa:
            acima_36 = sum(d["count"] for o, d in self.dados_pesquisa["idade"].items()
                          if any(x in o.lower() for x in ["36", "40", "50", "60", "mais"]))
            total = sum(d["count"] for d in self.dados_pesquisa["idade"].values())
            if total > 0:
                pct = int(round(acima_36 / total * 100, 0))
                media = f" (a média dos outros lançamento foi de {medias.get('acima_36_anos')}%)" if medias.get('acima_36_anos') else ""
                linhas.append(f"{pct}% acima de 36 anos{media}\n")

        # Estado civil (89% casadas)
        if "estado_civil" in self.dados_pesquisa:
            for opcao, dados in self.dados_pesquisa["estado_civil"].items():
                if "casado" in opcao.lower():
                    pct = int(round(dados["percentual"] * 100, 0))
                    media = f" (a média dos outros lançamento foi de {medias.get('casadas')}%)" if medias.get('casadas') else ""
                    linhas.append(f"{pct}% casadas{media}\n")

        # Escolaridade
        if "escolaridade" in self.dados_pesquisa:
            superior = sum(d["count"] for o, d in self.dados_pesquisa["escolaridade"].items()
                          if any(x in o.lower() for x in ["superior", "graduação", "mestrado", "doutorado"]))
            total = sum(d["count"] for d in self.dados_pesquisa["escolaridade"].values())
            if total > 0:
                pct = int(round(superior / total * 100, 0))
                media = f" (a média dos outros lançamento foi de {medias.get('superior_completo')}%)" if medias.get('superior_completo') else ""
                linhas.append(f"{pct}% com ao menos superior completo{media}\n")

        # Filhos
        if "filhos" in self.dados_pesquisa:
            ate_2 = sum(d["count"] for o, d in self.dados_pesquisa["filhos"].items()
                       if (o.strip().isdigit() and int(o.strip()) <= 2))
            total = sum(d["count"] for d in self.dados_pesquisa["filhos"].values())
            if total > 0:
                pct = int(round(ate_2 / total * 100, 0))
                media = f" (a média dos outros lançamento foi de {medias.get('ate_2_filhos')}%)" if medias.get('ate_2_filhos') else ""
                linhas.append(f"{pct}% com até 2 filhos{media}\n")

        # Religião
        if "religiao" in self.dados_pesquisa:
            for opcao, dados in self.dados_pesquisa["religiao"].items():
                if "católica" in opcao.lower():
                    pct = int(round(dados["percentual"] * 100, 0))
                    media = f" (a média dos outros lançamento foi de {medias.get('catolicas')}%)" if medias.get('catolicas') else ""
                    linhas.append(f"{pct}% católicas{media}\n")

        # Renda
        if "renda" in self.dados_pesquisa:
            acima_5k = sum(d["count"] for o, d in self.dados_pesquisa["renda"].items()
                          if any(x in o for x in ["R$ 5.000 -", "R$ 10.000", "+R$ 20.000"])
                          and "R$ 3.000 - R$ 5.000" not in o)
            total = sum(d["count"] for d in self.dados_pesquisa["renda"].values())
            if total > 0:
                pct = int(round(acima_5k / total * 100, 0))
                media = f" (a média dos outros lançamento foi de {medias.get('renda_acima_5k')}%)" if medias.get('renda_acima_5k') else ""
                linhas.append(f"{pct}% com renda acima de 5k{media}\n")

        return "".join(linhas)


def main():
    """Função principal"""
    parser = argparse.ArgumentParser(
        description="Gera resumo semanal de campanhas de marketing"
    )
    parser.add_argument("arquivo", help="Arquivo .xlsx exportado")
    parser.add_argument("data_inicio", help="Data inicial (DD/MM/YYYY)")
    parser.add_argument("data_fim", help="Data final (DD/MM/YYYY)")
    parser.add_argument("--historico", help="Arquivo JSON com histórico (opcional)")
    parser.add_argument("--output", help="Arquivo de saída (padrão: stdout)")

    args = parser.parse_args()

    # Cria instância e gera resumo
    resumo = ResumoSemanal(args.arquivo, args.data_inicio, args.data_fim, args.historico)
    texto = resumo.gerar_resumo()

    # Salva ou imprime
    if args.output:
        with open(args.output, 'w', encoding='utf-8') as f:
            f.write(texto)
        print(f"Resumo salvo em: {args.output}")
    else:
        print(texto)


if __name__ == "__main__":
    main()
